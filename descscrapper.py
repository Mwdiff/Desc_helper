#!C:\Users\hande\Documents\Python scripts\Desc_helper-master\desc_helper_venv\Scripts\python

from itertools import zip_longest
from pathlib import Path
from string import Template
from time import sleep

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

from secret import secret


def GetWebpage(url: str, s: requests.Session = requests.Session()) -> "Response":
    page = s.get(url)
    return page


def ExtractProductList(page: "Response") -> list:
    souped_page = BeautifulSoup(page.content, "lxml")
    products = souped_page.find_all("a", class_="product-name")
    product_list = [secret["site"] + product.get("href") for product in products]
    return product_list


def GetProductData(page: "Response") -> dict:
    souped_page = BeautifulSoup(page.content, "lxml")

    product_data = {"opis": [], "img": [], "titles": []}

    # Znajduje producenta produktu
    producent = souped_page.find("a", class_="firmlogo").contents[0].get("title")
    product_data["producent"] = producent

    # Znajduje SKU produktu
    sku = souped_page.find("div", itemprop="productID").text
    product_data["sku"] = sku

    # Cena zakupu netto
    net_price = souped_page.find("span", id="price_net_val")
    product_data["net"] = net_price.get_text().strip(" PLN").replace(" ", "")

    # Cena SRP
    srp_price = souped_page.find("span", id="srp_val")
    product_data["srp"] = srp_price.get_text().strip(" PLN").replace(" ", "")

    description = souped_page.find("div", id="component_projector_longdescription_not")

    # Tworzy listę nagłówków z opisu
    try:
        product_data["titles"] = [
            tit.get_text(strip=True)
            for tit in description.find_all("h3")
            if tit.get_text(strip=True)
        ]
        if not product_data["titles"]:
            product_data["titles"] = [
                tit.get_text(strip=True)
                for tit in description.find_all(style="font-size: 14pt;")
                if tit.get_text(strip=True)
            ]
    except:
        return product_data

    # Tworzy listę akapitów z opisu
    product_data["opis"] = [
        d.get_text(strip=True)
        for d in description.find_all(["p", "div"])
        if d.get_text(strip=True)
        and "iai_bottom" not in d.get_attribute_list("class")
        and "table-wrapper" not in d.get_attribute_list("class")
    ]
    
    try:
        if product_data["opis"][1] in product_data["opis"][0]:
            product_data["opis"].pop(0)
    except:
        pass

    product_data["opis"] = [
        item for item in product_data["opis"] if item not in product_data["titles"]
    ]

    # # test
    # try:
    #     print(product_data["opis"][0], end="\n\n")
    # except:
    #     print("err", end="\n\n")

    # Tworzy listę adresów zdjęć z opisu
    for img in description.find_all("img"):
        imgurl = img.get("src")
        if not "http" in imgurl:
            imgurl = secret["site"] + imgurl
        product_data["img"].append(imgurl)

    if len(product_data["img"]) > len(product_data["opis"]):
        nondupe = []
        [nondupe.append(i) for i in product_data["img"] if i not in nondupe]
        product_data["img"] = nondupe

    # Specyfikacja
    specification = description.find("tbody")
    try:
        spec = str(specification)
        reps = {
            "tbody": "ul",
            "tr>": "li>",
            "th>": "b>",
            "<br>": " ",
            "<td>": "",
            "</td>": "",
        }
        for o, n in reps.items():
            spec = spec.replace(o, n)

        product_data["spec"] = "<h2>Specyfikacja:</h2>\n" + spec
    except:
        product_data["spec"] = ""

    # W zestawie
    content = description.find("ul")
    if content != None:
        list = ""
        for sibling in content.previous_siblings:
            if sibling.name == "h3":
                header = sibling.get_text(strip=True)
                list = "<h2>" + header + ("" if ":" in header else ":") + "</h2>\n"
                break

        product_data["set"] = list + str(content)

    return product_data


def CompileDescription(data: dict) -> str:

    # szablon sekcji opisu
    Style_Sidetoside_head = '<section class="section">\n'
    Style_Sidetoside_template_1 = Template(
        """<div class="item item-6">
                <section class="image-item">
                    <img src="$img" />
                </section>
            </div>\n"""
    )

    Style_Sidetoside_template_2 = Template(
        """<div class="item item-6">
                <section class="text-item">
                    <h1>$title</h1><p>$section</p>
                </section>
            </div>\n"""
    )
    Style_Sidetoside_foot = "</section>\n"

    Style_Oneline_template_text = Template(
        """<section class="section">
                <div class="item item-12">
                    <section class="text-item">
                        <h2>$title</h2><p>$section</p>
                    </section>
                </div>
            </section>\n"""
    )

    Style_Oneline_template_img = Template(
        """<section class="section">
                <div class="item item-12">
                    <section class="image-item">
                        <img src="$img" />
                    </section>
                </div>
            </section>\n"""
    )

    # pierwsza sekcja specyfikacja + zawartość
    firstsection = data.setdefault("spec", "")

    if "set" in data:
        firstsection += "\n" + data["set"]

    if firstsection != "":
        desc = """<section class="section">
                        <div class="item item-12">
                            <section class="text-item">
                                {}
                            </section>
                        </div>
                    </section>\n""".format(
            firstsection
        )
    else:
        desc = ""

    # kolejne sekcje nagłówek+opis / zdjęcie naprzemiennie
    for t, (i, d) in zip(
        data["titles"], zip_longest(data["img"], data["opis"], fillvalue="")
    ):
        desc = (
            desc
            + Style_Oneline_template_text.substitute(img=i, title=t, section=d)
            + Style_Oneline_template_img.substitute(img=i, title=t, section=d)
        )
    return desc


def WriteToSheet(data: dict, file: str = "PlikDostawy") -> None:
    try:
        wb = load_workbook(f"./output/{file}.xlsx")
    except:
        wb = Workbook()
        ws = wb.create_sheet(file, 0)
        header = ("sku", "Zakup netto", "SRP", "Opis", "Zdjęcia ->")
        for i, tag in enumerate(header, start=1):
            ws.cell(1, i).value = tag
        Path("./output/").mkdir(exist_ok=True)

    ws = wb.active
    row = ws.max_row + 1

    desc = CompileDescription(data)

    ws.cell(row, 1).value = data["sku"]
    ws.cell(row, 2).value = data["net"]
    ws.cell(row, 3).value = data["srp"]
    ws.cell(row, 4).value = desc

    for i, img in enumerate(data["img"], start=5):
        ws.cell(row, i).value = img

    with open(f"./output/{file}_opis.txt", "a", encoding="utf8") as txt:
        print(
            "----------------------------------------------------------------------------------------------",
            file=txt,
        )
        print("SKU: " + data["sku"], file=txt, end="\n\n")
        print(desc, file=txt, end="\n\n")

    wb.save(f"./output/{file}.xlsx")


def ScrapeDelivery(
    URL: str, plik: str, s: requests.Session = requests.Session()
) -> None:
    p = s.post(secret["LOGIN_URL"], data=secret["payload"])

    products = ExtractProductList(GetWebpage(URL, s))

    for nr, product in enumerate(products, start=1):
        data = GetProductData(GetWebpage(product, s))
        WriteToSheet(data, plik)
        print("Zrobione: {}/{}".format(nr, len(products)), end="\t\t\r")
        sleep(0.5)


if __name__ == "__main__":

    with requests.Session() as session:
        p = session.post(secret["LOGIN_URL"], data=secret["payload"])
        
        while(session):
            URL = input("Podaj adres www dostawy: ")
            if not URL:
                break
            plik = input("Podaj nazwę pliku: ")
            
            ScrapeDelivery(URL, plik, session)
            print("\nGotowe!\n")
        
